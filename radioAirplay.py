import openpyxl as xl
from pandas import DataFrame, Series
import pandas as pd
import sys
import datetime

#
#    loadWorksheet: returns a frame, representing the Fan worksheet
#
def loadWorksheet():
    headers = {}
    fanInfo = {}

    wb = xl.load_workbook("/Users/gbradler/Documents/Songwriting/GregsSongs.xlsx")
    ws = wb.get_sheet_by_name("Radio Airplay")

    for row in ws.iter_rows('A1:J1'):
        for cell in row:
            headers[cell.column] = cell.value

    for row in ws.iter_rows('A2:J882'):
        fan = row[0]    
        info = {}
        for cell in row: 
            info[headers[cell.column]] = cell.value
        fanInfo[fan] = info

    frame = DataFrame(fanInfo).T
    return frame

#Main part of program ------------------------------------------------------
f = loadWorksheet()

if len(sys.argv) > 1:
    f = f[f['Fan Of'].str.startswith(sys.argv[1])]

print "\nTop stations:"
print f['Station'].value_counts().sort_values(ascending = False).head(50)
print "Number of distinct stations:", len(f['Station'].value_counts())

print "\nTop continents:"
print f['Continent'].value_counts().sort_values(ascending = False).head(3)

print "\nTop states:"
print f['State/Province'].value_counts().sort_values(ascending = False).head(50)
print "Number of distinct states:", len(f['State/Province'].value_counts())

print "\nGender breakdown:"
print f['Gender'].value_counts().sort_values(ascending = False)

print '\nMost fans per day:'
print f['Date'].value_counts().sort_values(ascending = False).head(5)

print '\nAge statistics:'
print 'Median fan age is: ' + str(f['Age'].replace('N/A',None).median())
print 'Average fan age is: ' , str(round(f['Age'].replace('N/A',None).mean()))

print "\nTop states/provinces:"
f['StateCountry'] = f['State/Province'] + "," +  f['Country']
print f['StateCountry'].value_counts().sort_values(ascending = False).head(3)

print "\nTop cities:"
g = f
g['CityCountry'] = f['City']  + "," +  g['Country']
print g['CityCountry'].value_counts().sort_values(ascending = False).head(3)

print "\nTop countries:"
print f['Country'].value_counts().sort_values(ascending = False).head(50)
print "Number of distinct countries:", len(f['Country'].value_counts())

print '\nDays of week stats:'
dow = Series([d.strftime("%A")  for d in f['Date']])
print dow.value_counts().sort_values(ascending=False)

print '\nBest months:'
monthYear = Series([d.strftime("%b%y")  for d in f['Date']])
print monthYear.value_counts().sort_values(ascending=False).head(7)

print "\nTotal count:"
print len(f)
